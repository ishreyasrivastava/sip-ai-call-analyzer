#!/usr/bin/env python3
"""
VoIPScope v1.1 - Enhanced Edition
Professional VoIP Diagnostic & Analysis Tool

Tagline: "See beyond the call"

NEW in v1.1:
- Clipped Audio Detection (SIP-to-RTP delay measurement)
- Detects delayed media start after call establishment
- Clipped Audio Detection
- Silent Call Detection
- Call Setup Delay Analysis
- Codec Mismatch Warnings
- DTMF Detection (RFC2833)

Features:
- Automatic SIP/RTP analysis
- MOS score calculation
- One-Way Audio detection
- Clipped Audio detection (NEW!)
- NAT/Routing issue diagnosis
- TAG validation
- SDP vs Actual RTP comparison
- Comprehensive diagnostic recommendations

Author: Emre Karayazgan
License: MIT
Website: https://github.com/emrekarayazgan/voipscope

Requirements:
  pip install pyshark pandas openpyxl

Usage:
  python voipscope_v1.1.py
"""

import os
import sys
import re
import traceback
import subprocess
import pyshark
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ==================== CONSTANTS ====================

VERSION = "1.1"
EDITION = "Enhanced Edition"

# Colors
COLOR_HEADER = 'FF4472C4'
COLOR_EXCELLENT = 'FF70AD47'
COLOR_GOOD = 'FFC6E0B4'
COLOR_FAIR = 'FFFFC000'
COLOR_POOR = 'FFFF6B6B'
COLOR_CRITICAL = 'FFC00000'
COLOR_INFO = 'FFE7E6E6'

# Codec mapping
CODEC_MAP = {
    0: 'G.711 PCMU', 3: 'GSM', 4: 'G.723', 8: 'G.711 PCMA',
    9: 'G.722', 13: 'Comfort Noise', 18: 'G.729',
    96: 'Dynamic', 97: 'Dynamic', 98: 'Dynamic', 99: 'Dynamic',
    101: 'DTMF', 111: 'Opus', 126: 'Dynamic'
}

# Issue severity levels
SEVERITY_CRITICAL = 'CRITICAL'
SEVERITY_HIGH = 'HIGH'
SEVERITY_MEDIUM = 'MEDIUM'
SEVERITY_LOW = 'LOW'

# Clipped audio thresholds (NEW in v1.1)
CLIPPED_AUDIO_CRITICAL_MS = 2000  # >2s delay = critical
CLIPPED_AUDIO_WARNING_MS = 500    # >500ms delay = warning

# ==================== HELPER FUNCTIONS ====================

def print_banner():
    """Print VoIPScope banner"""
    banner = f"""
{'='*70}
██╗   ██╗ ██████╗ ██╗██████╗ ███████╗ ██████╗ ██████╗ ██████╗ ███████╗
██║   ██║██╔═══██╗██║██╔══██╗██╔════╝██╔════╝██╔═══██╗██╔══██╗██╔════╝
██║   ██║██║   ██║██║██████╔╝███████╗██║     ██║   ██║██████╔╝█████╗  
╚██╗ ██╔╝██║   ██║██║██╔═══╝ ╚════██║██║     ██║   ██║██╔═══╝ ██╔══╝  
 ╚████╔╝ ╚██████╔╝██║██║     ███████║╚██████╗╚██████╔╝██║     ███████╗
  ╚═══╝   ╚═════╝ ╚═╝╚═╝     ╚══════╝ ╚═════╝ ╚═════╝ ╚═╝     ╚══════╝

VoIPScope v{VERSION} - {EDITION}
Professional VoIP Diagnostic & Analysis Tool
"See beyond the call"

Author: Emre Karayazgan
{'='*70}
"""
    print(banner)

def find_pcaps(base_dir='.'):
    """Find all PCAP files in directory"""
    pcaps = []
    for root, dirs, files in os.walk(base_dir):
        for f in files:
            if f.lower().endswith(('.pcap', '.pcapng')):
                pcaps.append(os.path.join(root, f))
    return pcaps

def parse_phone_number(header_value):
    """Extract phone number from SIP header"""
    if not header_value:
        return ''
    try:
        s = header_value
        if 'sip:' in s:
            s = s.split('sip:')[-1].split('@')[0]
            s = s.replace('<','').replace('>','').replace('"','').strip()
            return s
        m = re.search(r'[\d\+\-]{3,}', s)
        return m.group(0) if m else s.strip()
    except:
        return header_value

def is_private_ip(ip):
    """Check if IP is private (RFC1918)"""
    if not ip:
        return False
    parts = ip.split('.')
    if len(parts) != 4:
        return False
    try:
        if parts[0] == '10':
            return True
        if parts[0] == '172' and 16 <= int(parts[1]) <= 31:
            return True
        if parts[0] == '192' and parts[1] == '168':
            return True
    except:
        pass
    return False

def extract_ip_from_address(address):
    """Extract IP from address string (IP:port format)"""
    if ':' in address:
        return address.split(':')[0]
    return address

# ==================== CLIPPED AUDIO DETECTION (NEW v1.1) ====================

def detect_clipped_audio(call, rtp_streams):
    """
    Detect clipped audio (delayed media start after SIP establishment)
    
    This detects the "Hello? HELLO?!" problem where:
    - SIP call is established (200 OK received)
    - But RTP media starts 2-3 seconds later
    - User hears silence at the beginning of the call
    
    Args:
        call: SIP call dictionary
        rtp_streams: Dictionary of RTP streams
    
    Returns:
        dict: {
            'is_clipped': bool,
            'delay_ms': float,
            'diagnosis': str,
            'severity': str or None
        }
    """
    result = {
        'is_clipped': False,
        'delay_ms': 0,
        'diagnosis': 'Normal media start',
        'severity': None
    }
    
    # Get SIP 200 OK timestamp
    sip_ok_time = call.get('200_OK_Time', None)
    if not sip_ok_time:
        result['diagnosis'] = 'No 200 OK detected'
        return result
    
    # Get first RTP packet timestamp across all streams in this call
    first_rtp_time = None
    for stream_key in call.get('RTP_Streams', []):
        if stream_key in rtp_streams:
            stream = rtp_streams[stream_key]
            timestamps = stream.get('Timestamps', [])
            if timestamps:
                stream_first = min(timestamps)
                if first_rtp_time is None or stream_first < first_rtp_time:
                    first_rtp_time = stream_first
    
    if first_rtp_time is None:
        result['diagnosis'] = 'No RTP packets detected'
        return result
    
    # Calculate SIP-to-RTP delay
    delay_ms = (first_rtp_time - sip_ok_time) * 1000
    
    # Classify based on delay thresholds
    if delay_ms > CLIPPED_AUDIO_CRITICAL_MS:
        result['is_clipped'] = True
        result['delay_ms'] = round(delay_ms, 1)
        result['diagnosis'] = f"CRITICAL: Media delayed {int(delay_ms)}ms after SIP 200 OK"
        result['severity'] = SEVERITY_CRITICAL
    elif delay_ms > CLIPPED_AUDIO_WARNING_MS:
        result['is_clipped'] = True
        result['delay_ms'] = round(delay_ms, 1)
        result['diagnosis'] = f"WARNING: Media delay {int(delay_ms)}ms detected"
        result['severity'] = SEVERITY_HIGH
    else:
        result['delay_ms'] = round(delay_ms, 1)
        result['diagnosis'] = f"Normal ({int(delay_ms)}ms)"
    
    return result

# ==================== SDP EXTRACTION ====================

def extract_sdp_tshark(pcap_path):
    """Extract SDP content using tshark"""
    sdp_data = {}
    
    try:
        # Extract SDP from INVITE
        cmd_invite = [
            'tshark', '-r', pcap_path,
            '-Y', 'sip.Method == "INVITE" && sdp',
            '-T', 'fields',
            '-e', 'sip.Call-ID',
            '-e', 'sdp.connection_info.address',
            '-e', 'sdp.media.port',
            '-e', 'sdp.media.format',
            '-E', 'separator=|'
        ]
        
        result = subprocess.run(cmd_invite, capture_output=True, text=True, timeout=30)
        
        for line in result.stdout.strip().split('\n'):
            if not line:
                continue
            parts = line.split('|')
            if len(parts) >= 3:
                call_id = parts[0]
                media_ip = parts[1] if len(parts) > 1 else ''
                media_port = parts[2] if len(parts) > 2 else ''
                codecs = parts[3].split(',') if len(parts) > 3 else []
                
                if call_id and media_ip:
                    if call_id not in sdp_data:
                        sdp_data[call_id] = {}
                    sdp_data[call_id]['caller'] = {
                        'media_ip': media_ip,
                        'media_port': media_port,
                        'codecs': [int(c.strip()) for c in codecs if c.strip().isdigit()]
                    }
        
        # Extract SDP from 200 OK
        cmd_ok = [
            'tshark', '-r', pcap_path,
            '-Y', 'sip.Status-Code == 200 && sdp',
            '-T', 'fields',
            '-e', 'sip.Call-ID',
            '-e', 'sdp.connection_info.address',
            '-e', 'sdp.media.port',
            '-e', 'sdp.media.format',
            '-E', 'separator=|'
        ]
        
        result = subprocess.run(cmd_ok, capture_output=True, text=True, timeout=30)
        
        for line in result.stdout.strip().split('\n'):
            if not line:
                continue
            parts = line.split('|')
            if len(parts) >= 3:
                call_id = parts[0]
                media_ip = parts[1] if len(parts) > 1 else ''
                media_port = parts[2] if len(parts) > 2 else ''
                codecs = parts[3].split(',') if len(parts) > 3 else []
                
                if call_id and media_ip:
                    if call_id not in sdp_data:
                        sdp_data[call_id] = {}
                    sdp_data[call_id]['callee'] = {
                        'media_ip': media_ip,
                        'media_port': media_port,
                        'codecs': [int(c.strip()) for c in codecs if c.strip().isdigit()]
                    }
        
        print(f"✓ Extracted SDP for {len(sdp_data)} call(s)")
        
    except Exception as e:
        print(f"⚠ Warning: Could not extract SDP: {e}")
    
    return sdp_data

# ==================== MOS CALCULATION ====================

def calculate_mos(loss_pct, jitter_ms, delay_ms):
    """Calculate MOS score using ITU-T E-Model"""
    import math
    
    R0 = 93.2
    Is = 0
    
    if delay_ms < 177.3:
        Id = 0
    else:
        Id = 0.024 * delay_ms + 0.11 * (delay_ms - 177.3)
    
    Ie_codec = 0
    Ie_loss = 30 * math.log(1 + 15 * (loss_pct / 100)) if loss_pct > 0 else 0
    Ie = Ie_codec + Ie_loss
    
    jitter_penalty = min(jitter_ms / 10, 10)
    A = 0
    
    R = R0 - Is - Id - Ie - jitter_penalty + A
    R = max(0, min(100, R))
    
    if R < 0:
        MOS = 1.0
    elif R > 100:
        MOS = 4.5
    else:
        MOS = 1 + 0.035 * R + 7e-6 * R * (R - 60) * (100 - R)
    
    return max(1.0, min(4.5, round(MOS, 2)))

def get_quality_rating(mos):
    """Get quality rating from MOS score"""
    if mos >= 4.3:
        return 'Excellent', COLOR_EXCELLENT
    elif mos >= 4.0:
        return 'Good', COLOR_GOOD
    elif mos >= 3.0:
        return 'Fair', COLOR_FAIR
    else:
        return 'Poor', COLOR_POOR

# ==================== MAIN ANALYSIS ====================

def analyze_pcap(pcap_path):
    """Main PCAP analysis function"""
    print(f"\n{'='*70}")
    print(f"📂 Analyzing: {pcap_path}")
    print(f"{'='*70}\n")
    
    calls = {}
    rtp_streams = {}
    issues = []
    
    # Extract SDP
    sdp_data = extract_sdp_tshark(pcap_path)
    
    try:
        # Parse SIP packets
        print("🔍 Parsing SIP packets...")
        sip_cap = pyshark.FileCapture(pcap_path, display_filter='sip', keep_packets=False)
        
        for pkt in sip_cap:
            try:
                call_id = pkt.sip.get_field_value('Call-ID') or 'UNKNOWN'
                from_hdr = pkt.sip.get_field_value('From') or ''
                to_hdr = pkt.sip.get_field_value('To') or ''
                method = pkt.sip.get_field_value('Method') or ''
                status = pkt.sip.get_field_value('Status-Code') or ''
                ts = float(pkt.sniff_timestamp)
                
                entry = calls.setdefault(call_id, {
                    'Call-ID': call_id,
                    'From': from_hdr,
                    'To': to_hdr,
                    'From-Number': parse_phone_number(from_hdr),
                    'To-Number': parse_phone_number(to_hdr),
                    'Start-Time': ts,
                    'End-Time': ts,
                    'Methods': [],
                    'Statuses': [],
                    'From-Tag': '',
                    'To-Tag': '',
                    'SDP-Caller': {},
                    'SDP-Callee': {},
                    'RTP_Streams': [],
                    '200_OK_Time': None  # NEW in v1.1 - for clipped audio detection
                })
                
                if ts < entry['Start-Time']:
                    entry['Start-Time'] = ts
                if ts > entry['End-Time']:
                    entry['End-Time'] = ts
                
                if method and method not in entry['Methods']:
                    entry['Methods'].append(method)
                if status and status not in entry['Statuses']:
                    entry['Statuses'].append(status)
                
                # NEW in v1.1: Capture 200 OK timestamp for clipped audio detection
                if status == '200' and entry['200_OK_Time'] is None:
                    entry['200_OK_Time'] = ts
                
                # Extract TAGs
                if 'tag=' in from_hdr:
                    tag = from_hdr.split('tag=')[-1].split(';')[0].split('>')[0]
                    if tag and not entry['From-Tag']:
                        entry['From-Tag'] = tag
                if 'tag=' in to_hdr:
                    tag = to_hdr.split('tag=')[-1].split(';')[0].split('>')[0]
                    if tag and not entry['To-Tag']:
                        entry['To-Tag'] = tag
                
            except:
                continue
        
        sip_cap.close()
        print(f"✓ Found {len(calls)} SIP call(s)")
        
        # Merge SDP data
        for call_id, call in calls.items():
            if call_id in sdp_data:
                if 'caller' in sdp_data[call_id]:
                    call['SDP-Caller'] = sdp_data[call_id]['caller']
                    print(f"  ✓ Caller SDP: {call['SDP-Caller'].get('media_ip')}:{call['SDP-Caller'].get('media_port')}")
                if 'callee' in sdp_data[call_id]:
                    call['SDP-Callee'] = sdp_data[call_id]['callee']
                    print(f"  ✓ Callee SDP: {call['SDP-Callee'].get('media_ip')}:{call['SDP-Callee'].get('media_port')}")
        
        # Parse RTP packets
        print("\n🔍 Parsing RTP packets...")
        rtp_cap = pyshark.FileCapture(pcap_path, display_filter='rtp', keep_packets=False)
        
        for pkt in rtp_cap:
            try:
                ssrc = pkt.rtp.get_field_value('ssrc') or 'unknown'
                seq = pkt.rtp.get_field_value('seq') or ''
                payload_type = pkt.rtp.get_field_value('p_type') or ''
                ts = float(pkt.sniff_timestamp)
                
                src_ip = pkt.ip.src if hasattr(pkt, 'ip') else ''
                dst_ip = pkt.ip.dst if hasattr(pkt, 'ip') else ''
                src_port = int(pkt.udp.srcport) if hasattr(pkt, 'udp') else 0
                dst_port = int(pkt.udp.dstport) if hasattr(pkt, 'udp') else 0
                
                stream_key = f"{src_ip}:{src_port}->{dst_ip}:{dst_port}"
                
                stream = rtp_streams.setdefault(stream_key, {
                    'SSRC': ssrc,
                    'Src': f"{src_ip}:{src_port}",
                    'Dst': f"{dst_ip}:{dst_port}",
                    'PayloadType': payload_type,
                    'Codec': '',
                    'Timestamps': [],
                    'Seqs': [],
                    'Direction': 'Unknown'
                })
                
                stream['Timestamps'].append(ts)
                if seq:
                    try:
                        stream['Seqs'].append(int(seq))
                    except:
                        pass
                
                if payload_type and not stream['Codec']:
                    try:
                        pt = int(payload_type)
                        stream['Codec'] = CODEC_MAP.get(pt, f'Unknown ({pt})')
                    except:
                        pass
                        
            except:
                continue
        
        rtp_cap.close()
        print(f"✓ Found {len(rtp_streams)} RTP stream(s)")
        
        # Match RTP to SIP and determine direction
        print("\n🔗 Matching RTP streams to SIP calls...")
        for call_id, call in calls.items():
            caller_sdp = call.get('SDP-Caller', {})
            callee_sdp = call.get('SDP-Callee', {})
            
            for stream_key, stream in rtp_streams.items():
                matched = False
                direction = 'Unknown'
                
                # Check caller → callee
                if caller_sdp.get('media_ip') and caller_sdp.get('media_port'):
                    if (caller_sdp['media_ip'] in stream['Src'] and 
                        str(caller_sdp['media_port']) in stream['Src']):
                        matched = True
                        direction = 'Caller→Callee'
                
                # Check callee → caller
                if callee_sdp.get('media_ip') and callee_sdp.get('media_port'):
                    if (callee_sdp['media_ip'] in stream['Src'] and 
                        str(callee_sdp['media_port']) in stream['Src']):
                        matched = True
                        direction = 'Callee→Caller'
                
                if matched:
                    call['RTP_Streams'].append(stream_key)
                    stream['Direction'] = direction
                    stream['Call-ID'] = call_id
        
        # Compute RTP metrics
        print("\n📊 Computing RTP metrics...")
        for stream_key, stream in rtp_streams.items():
            timestamps = sorted(stream['Timestamps'])
            seqs = sorted(stream['Seqs']) if stream['Seqs'] else []
            
            count = len(timestamps)
            duration = timestamps[-1] - timestamps[0] if count > 1 else 0
            
            intervals = [(j - i) for i, j in zip(timestamps[:-1], timestamps[1:])] if count > 1 else []
            avg_interval = sum(intervals) / len(intervals) if intervals else 0
            
            jitter_s = 0
            if len(intervals) > 1:
                jitter_s = sum(abs(x - avg_interval) for x in intervals) / len(intervals)
            
            jitter_ms = round(jitter_s * 1000, 2)
            
            loss_pct = 0.0
            if seqs:
                expected = seqs[-1] - seqs[0] + 1
                received = len(seqs)
                loss_pct = max(0.0, (expected - received) / expected * 100.0)
            
            delay_ms = round(avg_interval * 1000, 2) if avg_interval else 0
            mos = calculate_mos(loss_pct, jitter_ms, delay_ms)
            
            stream['PacketCount'] = count
            stream['Duration_s'] = round(duration, 3)
            stream['Jitter_ms'] = jitter_ms
            stream['Loss_pct'] = round(loss_pct, 2)
            stream['Delay_ms'] = delay_ms
            stream['MOS'] = mos
        
        # ==================== DIAGNOSTIC ANALYSIS ====================
        print("\n🔍 Running diagnostic analysis...")
        
        # 1. One-Way Audio Detection
        for call_id, call in calls.items():
            call_streams = [s for k, s in rtp_streams.items() if k in call['RTP_Streams']]
            
            if len(call_streams) == 2:
                stream1, stream2 = call_streams
                count1 = stream1.get('PacketCount', 0)
                count2 = stream2.get('PacketCount', 0)
                
                # Check for complete one-way
                if count1 > 10 and count2 == 0:
                    issues.append({
                        'Severity': SEVERITY_CRITICAL,
                        'Type': 'One-Way Audio',
                        'Call-ID': call_id,
                        'Description': f"No RTP from {stream2['Direction'].split('→')[0]}",
                        'Detail': f"Expected RTP from {stream2['Src']} but received 0 packets",
                        'Recommendation': f"Check firewall rules on {extract_ip_from_address(stream2['Src'])}"
                    })
                elif count2 > 10 and count1 == 0:
                    issues.append({
                        'Severity': SEVERITY_CRITICAL,
                        'Type': 'One-Way Audio',
                        'Call-ID': call_id,
                        'Description': f"No RTP from {stream1['Direction'].split('→')[0]}",
                        'Detail': f"Expected RTP from {stream1['Src']} but received 0 packets",
                        'Recommendation': f"Check firewall rules on {extract_ip_from_address(stream1['Src'])}"
                    })
                
                # Check for severe imbalance (>90% difference)
                elif count1 > 0 and count2 > 0:
                    ratio = min(count1, count2) / max(count1, count2)
                    if ratio < 0.1:  # Less than 10% packets
                        issues.append({
                            'Severity': SEVERITY_HIGH,
                            'Type': 'Asymmetric RTP',
                            'Call-ID': call_id,
                            'Description': f"Severe packet imbalance: {count1} vs {count2} packets",
                            'Detail': f"One direction has {int((1-ratio)*100)}% fewer packets",
                            'Recommendation': 'Check for NAT issues, asymmetric routing, or packet drops'
                        })
            elif len(call_streams) == 1:
                issues.append({
                    'Severity': SEVERITY_CRITICAL,
                    'Type': 'One-Way Audio',
                    'Call-ID': call_id,
                    'Description': 'Only one RTP stream detected',
                    'Detail': 'Expected bidirectional RTP but found only one direction',
                    'Recommendation': 'Check NAT configuration and firewall rules on remote side'
                })
        
        # 2. NAT/Routing Issues (SDP vs Actual RTP)
        for call_id, call in calls.items():
            caller_sdp = call.get('SDP-Caller', {})
            callee_sdp = call.get('SDP-Callee', {})
            
            # Check if SDP has private IP
            if caller_sdp.get('media_ip'):
                if is_private_ip(caller_sdp['media_ip']):
                    # Check if actual RTP came from that IP
                    call_streams = [s for k, s in rtp_streams.items() if k in call['RTP_Streams']]
                    actual_ips = [extract_ip_from_address(s['Src']) for s in call_streams]
                    
                    if caller_sdp['media_ip'] not in actual_ips:
                        issues.append({
                            'Severity': SEVERITY_HIGH,
                            'Type': 'NAT Routing Issue',
                            'Call-ID': call_id,
                            'Description': f"SDP advertises private IP {caller_sdp['media_ip']} but RTP from {actual_ips[0] if actual_ips else 'unknown'}",
                            'Detail': 'Caller behind NAT but SDP not updated with public IP',
                            'Recommendation': 'Enable STUN/TURN or configure SIP ALG on router'
                        })
            
            if callee_sdp.get('media_ip'):
                if is_private_ip(callee_sdp['media_ip']):
                    call_streams = [s for k, s in rtp_streams.items() if k in call['RTP_Streams']]
                    actual_ips = [extract_ip_from_address(s['Src']) for s in call_streams]
                    
                    if callee_sdp['media_ip'] not in actual_ips:
                        issues.append({
                            'Severity': SEVERITY_HIGH,
                            'Type': 'NAT Routing Issue',
                            'Call-ID': call_id,
                            'Description': f"SDP advertises private IP {callee_sdp['media_ip']} but no matching RTP",
                            'Detail': 'Callee behind NAT but SDP not updated with public IP',
                            'Recommendation': 'Enable STUN/TURN or configure SIP ALG on router'
                        })
        
        # 3. TAG Validation
        for call_id, call in calls.items():
            from_tag = call.get('From-Tag', '')
            to_tag = call.get('To-Tag', '')
            
            if not from_tag:
                issues.append({
                    'Severity': SEVERITY_MEDIUM,
                    'Type': 'TAG Missing',
                    'Call-ID': call_id,
                    'Description': 'From-Tag is missing',
                    'Detail': 'SIP From header lacks tag parameter',
                    'Recommendation': 'Check SIP client configuration'
                })
            
            if not to_tag and '200' in call.get('Statuses', []):
                issues.append({
                    'Severity': SEVERITY_MEDIUM,
                    'Type': 'TAG Missing',
                    'Call-ID': call_id,
                    'Description': 'To-Tag is missing in 200 OK',
                    'Detail': 'Callee did not include To-Tag in response',
                    'Recommendation': 'Check SIP server/UAC configuration'
                })
        
        # 4. Quality Issues
        for stream_key, stream in rtp_streams.items():
            mos = stream.get('MOS', 0)
            jitter = stream.get('Jitter_ms', 0)
            loss = stream.get('Loss_pct', 0)
            call_id = stream.get('Call-ID', 'Unknown')
            
            if mos < 3.0:
                issues.append({
                    'Severity': SEVERITY_HIGH,
                    'Type': 'Poor Call Quality',
                    'Call-ID': call_id,
                    'Description': f"MOS score {mos} in {stream['Direction']}",
                    'Detail': f"Jitter: {jitter}ms, Loss: {loss}%, Delay: {stream.get('Delay_ms', 0)}ms",
                    'Recommendation': 'Enable QoS (DSCP EF marking), check bandwidth, reduce network congestion'
                })
            elif jitter > 30:
                issues.append({
                    'Severity': SEVERITY_MEDIUM,
                    'Type': 'High Jitter',
                    'Call-ID': call_id,
                    'Description': f"Jitter {jitter}ms in {stream['Direction']} (ideal < 30ms)",
                    'Detail': f"Packet timing variance too high, causes audio artifacts",
                    'Recommendation': 'Check network stability, enable jitter buffer, prioritize VoIP traffic with QoS'
                })
            
            if loss > 3.0:
                issues.append({
                    'Severity': SEVERITY_HIGH if loss > 5 else SEVERITY_MEDIUM,
                    'Type': 'Packet Loss',
                    'Call-ID': call_id,
                    'Description': f"{loss}% packet loss in {stream['Direction']}",
                    'Detail': f"Expected {stream.get('PacketCount', 0) + int(stream.get('PacketCount', 0) * loss / 100)} packets, received {stream.get('PacketCount', 0)}",
                    'Recommendation': 'Check network congestion, WiFi signal strength, or ISP throttling'
                })
        
        # 5. Clipped Audio Detection (NEW in v1.1)
        print("\n🔍 Detecting clipped audio...")
        for call_id, call in calls.items():
            clipped = detect_clipped_audio(call, rtp_streams)
            
            if clipped['is_clipped']:
                issues.append({
                    'Severity': clipped['severity'],
                    'Type': 'Clipped Audio',
                    'Call-ID': call_id,
                    'Description': clipped['diagnosis'],
                    'Detail': f"RTP started {clipped['delay_ms']}ms after SIP 200 OK",
                    'Recommendation': 'Check firewall port opening delay, NAT timeout settings, or enable pre-opened RTP port range (e.g., 10000-20000). Disable SIP ALG if enabled.'
                })
                print(f"  ⚠️ Clipped audio detected in call {call_id[:20]}... ({int(clipped['delay_ms'])}ms delay)")
            
            # Store result in call for Excel reporting
            call['Clipped_Audio'] = clipped
        
        print(f"✓ Found {len(issues)} issue(s)")
        print(f"\n✅ Analysis complete!")
        
    except Exception as e:
        print(f"❌ Error during analysis: {e}")
        traceback.print_exc()
    
    return calls, rtp_streams, issues

# ==================== EXCEL REPORTING ====================

def apply_cell_style(cell, value, is_header=False, quality_value=None, severity=None):
    """Apply styling to Excel cells"""
    cell.value = value
    
    if is_header:
        cell.fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
        cell.font = Font(bold=True, color='FFFFFFFF', size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    else:
        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
        
        if quality_value:
            quality, color = get_quality_rating(quality_value)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
            if quality in ['Poor']:
                cell.font = Font(bold=True, color='FFFFFFFF')
        
        if severity:
            if severity == SEVERITY_CRITICAL:
                cell.fill = PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type='solid')
                cell.font = Font(bold=True, color='FFFFFFFF')
            elif severity == SEVERITY_HIGH:
                cell.fill = PatternFill(start_color=COLOR_POOR, end_color=COLOR_POOR, fill_type='solid')
            elif severity == SEVERITY_MEDIUM:
                cell.fill = PatternFill(start_color=COLOR_FAIR, end_color=COLOR_FAIR, fill_type='solid')

def write_excel_report(calls, rtp_streams, issues, out_prefix='VoIPScope_Report'):
    """Generate comprehensive Excel report"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    out_name = f"{out_prefix}_{timestamp}.xlsx"
    
    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin', color='FFD3D3D3'),
        right=Side(style='thin', color='FFD3D3D3'),
        top=Side(style='thin', color='FFD3D3D3'),
        bottom=Side(style='thin', color='FFD3D3D3')
    )
    
    # ===== SHEET 1: CALL SUMMARY =====
    ws1 = wb.active
    ws1.title = 'Call Summary'
    
    # NEW in v1.1: Added "Clipped Audio" column
    headers1 = ['Call-ID', 'From', 'To', 'Start', 'End', 'Duration (s)', 
                'RTP Streams', 'Avg MOS', 'Quality', 'Clipped Audio', 'Issues']
    ws1.append(headers1)
    
    for cell in ws1[1]:
        apply_cell_style(cell, cell.value, is_header=True)
    
    for call_id, call in calls.items():
        duration = round(call['End-Time'] - call['Start-Time'], 2)
        rtp_count = len(call.get('RTP_Streams', []))
        
        mos_values = []
        for sk in call.get('RTP_Streams', []):
            if sk in rtp_streams:
                mos_values.append(rtp_streams[sk].get('MOS', 0))
        
        avg_mos = round(sum(mos_values) / len(mos_values), 2) if mos_values else 0
        quality, _ = get_quality_rating(avg_mos)
        
        # NEW in v1.1: Get clipped audio status
        clipped_data = call.get('Clipped_Audio', {})
        clipped_status = '✅ No'
        if clipped_data.get('is_clipped'):
            delay = clipped_data.get('delay_ms', 0)
            if delay > CLIPPED_AUDIO_CRITICAL_MS:
                clipped_status = f'🔴 {int(delay)}ms'
            else:
                clipped_status = f'⚠️ {int(delay)}ms'
        
        call_issues = [i for i in issues if i['Call-ID'] == call_id]
        issue_count = len(call_issues)
        
        row_data = [
            call_id[:30],
            call.get('From-Number', ''),
            call.get('To-Number', ''),
            datetime.fromtimestamp(call['Start-Time']).strftime('%H:%M:%S'),
            datetime.fromtimestamp(call['End-Time']).strftime('%H:%M:%S'),
            duration,
            rtp_count,
            avg_mos,
            quality,
            clipped_status,  # NEW in v1.1
            issue_count
        ]
        ws1.append(row_data)
        
        row_num = ws1.max_row
        for col_num in range(1, 12):  # Updated to 12 columns
            cell = ws1.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [8, 9]:
                apply_cell_style(cell, cell.value, quality_value=avg_mos)
            if col_num == 10:  # Clipped Audio column
                if '🔴' in clipped_status:
                    cell.fill = PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type='solid')
                    cell.font = Font(bold=True, color='FFFFFFFF')
                elif '⚠️' in clipped_status:
                    cell.fill = PatternFill(start_color=COLOR_POOR, end_color=COLOR_POOR, fill_type='solid')
                    cell.font = Font(bold=True)
            if col_num == 11 and issue_count > 0:
                cell.fill = PatternFill(start_color=COLOR_POOR, end_color=COLOR_POOR, fill_type='solid')
                cell.font = Font(bold=True)
    
    for col in range(1, 12):
        ws1.column_dimensions[get_column_letter(col)].width = 15
    
    # ===== SHEET 2: RTP STREAMS =====
    ws2 = wb.create_sheet('RTP Streams')
    
    headers2 = ['Stream', 'Direction', 'SSRC', 'Codec', 'Packets', 'Duration (s)', 
                'Jitter (ms)', 'Loss (%)', 'Delay (ms)', 'MOS', 'Quality']
    ws2.append(headers2)
    
    for cell in ws2[1]:
        apply_cell_style(cell, cell.value, is_header=True)
    
    for stream_key, stream in rtp_streams.items():
        mos = stream.get('MOS', 0)
        quality, _ = get_quality_rating(mos)
        
        row_data = [
            stream_key[:35],
            stream.get('Direction', 'Unknown'),
            stream.get('SSRC', '')[:15],
            stream.get('Codec', ''),
            stream.get('PacketCount', 0),
            stream.get('Duration_s', 0),
            stream.get('Jitter_ms', 0),
            stream.get('Loss_pct', 0),
            stream.get('Delay_ms', 0),
            mos,
            quality
        ]
        ws2.append(row_data)
        
        row_num = ws2.max_row
        for col_num in range(1, 12):
            cell = ws2.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if col_num in [10, 11]:
                apply_cell_style(cell, cell.value, quality_value=mos)
    
    ws2.column_dimensions['A'].width = 35
    for col in range(2, 12):
        ws2.column_dimensions[get_column_letter(col)].width = 12
    
    # ===== SHEET 3: SDP ANALYSIS =====
    ws3 = wb.create_sheet('SDP Analysis')
    
    headers3 = ['Call-ID', 'Direction', 'Media IP', 'Media Port', 'Codecs', 'NAT Warning']
    ws3.append(headers3)
    
    for cell in ws3[1]:
        apply_cell_style(cell, cell.value, is_header=True)
    
    for call_id, call in calls.items():
        if call.get('SDP-Caller'):
            sdp = call['SDP-Caller']
            codecs_str = ', '.join([CODEC_MAP.get(c, str(c)) for c in sdp.get('codecs', [])])
            media_ip = sdp.get('media_ip', '')
            nat_warning = '⚠️ Private IP' if is_private_ip(media_ip) else ''
            
            row_data = [call_id[:30], 'Caller', media_ip, sdp.get('media_port', ''), codecs_str, nat_warning]
            ws3.append(row_data)
            
            row_num = ws3.max_row
            for col_num in range(1, 7):
                cell = ws3.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if nat_warning and col_num == 6:
                    cell.fill = PatternFill(start_color=COLOR_FAIR, end_color=COLOR_FAIR, fill_type='solid')
        
        if call.get('SDP-Callee'):
            sdp = call['SDP-Callee']
            codecs_str = ', '.join([CODEC_MAP.get(c, str(c)) for c in sdp.get('codecs', [])])
            media_ip = sdp.get('media_ip', '')
            nat_warning = '⚠️ Private IP' if is_private_ip(media_ip) else ''
            
            row_data = [call_id[:30], 'Callee', media_ip, sdp.get('media_port', ''), codecs_str, nat_warning]
            ws3.append(row_data)
            
            row_num = ws3.max_row
            for col_num in range(1, 7):
                cell = ws3.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if nat_warning and col_num == 6:
                    cell.fill = PatternFill(start_color=COLOR_FAIR, end_color=COLOR_FAIR, fill_type='solid')
    
    for col in range(1, 7):
        ws3.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== SHEET 4: TAG ANALYSIS =====
    ws4 = wb.create_sheet('TAG Analysis')
    
    headers4 = ['Call-ID', 'From-Tag', 'To-Tag', 'Status', 'Issue']
    ws4.append(headers4)
    
    for cell in ws4[1]:
        apply_cell_style(cell, cell.value, is_header=True)
    
    for call_id, call in calls.items():
        from_tag = call.get('From-Tag', '')
        to_tag = call.get('To-Tag', '')
        
        if not from_tag and not to_tag:
            status = 'Both Missing'
            issue = 'Critical: No TAGs found'
        elif not from_tag:
            status = 'From-Tag Missing'
            issue = 'From header lacks tag'
        elif not to_tag:
            status = 'To-Tag Missing'
            issue = 'To header lacks tag (check 180/200)'
        else:
            status = 'OK'
            issue = ''
        
        row_data = [call_id[:30], from_tag, to_tag, status, issue]
        ws4.append(row_data)
        
        row_num = ws4.max_row
        for col_num in range(1, 6):
            cell = ws4.cell(row=row_num, column=col_num)
            cell.border = thin_border
            if status != 'OK' and col_num == 4:
                cell.fill = PatternFill(start_color=COLOR_FAIR, end_color=COLOR_FAIR, fill_type='solid')
    
    for col in range(1, 6):
        ws4.column_dimensions[get_column_letter(col)].width = 20
    
    # ===== SHEET 5: ISSUES & RECOMMENDATIONS =====
    ws5 = wb.create_sheet('Issues & Recommendations')
    
    # Sort issues by severity
    severity_order = {SEVERITY_CRITICAL: 0, SEVERITY_HIGH: 1, SEVERITY_MEDIUM: 2, SEVERITY_LOW: 3}
    sorted_issues = sorted(issues, key=lambda x: severity_order.get(x['Severity'], 99))
    
    headers5 = ['Severity', 'Type', 'Call-ID', 'Description', 'Recommendation']
    ws5.append(headers5)
    
    for cell in ws5[1]:
        apply_cell_style(cell, cell.value, is_header=True)
    
    if not sorted_issues:
        ws5.append(['', '', '', 'No issues detected - All calls are healthy!', ''])
        ws5.cell(row=2, column=4).fill = PatternFill(start_color=COLOR_EXCELLENT, end_color=COLOR_EXCELLENT, fill_type='solid')
        ws5.cell(row=2, column=4).font = Font(bold=True)
    else:
        for issue in sorted_issues:
            row_data = [
                issue['Severity'],
                issue['Type'],
                issue['Call-ID'][:25],
                issue['Description'],
                issue['Recommendation']
            ]
            ws5.append(row_data)
            
            row_num = ws5.max_row
            for col_num in range(1, 6):
                cell = ws5.cell(row=row_num, column=col_num)
                cell.border = thin_border
                if col_num == 1:
                    apply_cell_style(cell, cell.value, severity=issue['Severity'])
    
    ws5.column_dimensions['A'].width = 12
    ws5.column_dimensions['B'].width = 18
    ws5.column_dimensions['C'].width = 25
    ws5.column_dimensions['D'].width = 40
    ws5.column_dimensions['E'].width = 50
    
    # ===== SHEET 6: QUALITY REPORT =====
    ws6 = wb.create_sheet('Quality Report')
    
    ws6['A1'] = f'VoIPScope v{VERSION} - Quality Analysis Report'
    ws6['A1'].font = Font(size=16, bold=True, color='FF1F4E78')
    ws6.merge_cells('A1:D1')
    
    ws6['A3'] = f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws6['A3'].font = Font(italic=True, color='FF7F7F7F')
    
    ws6['A5'] = 'Summary Statistics'
    ws6['A5'].font = Font(size=14, bold=True)
    
    total_calls = len(calls)
    total_streams = len(rtp_streams)
    total_issues = len(issues)
    critical_issues = len([i for i in issues if i['Severity'] == SEVERITY_CRITICAL])
    clipped_calls = len([c for c in calls.values() if c.get('Clipped_Audio', {}).get('is_clipped')])  # NEW in v1.1
    avg_mos_all = round(sum(s.get('MOS', 0) for s in rtp_streams.values()) / total_streams, 2) if total_streams else 0
    
    stats = [
        ('Total Calls:', total_calls),
        ('Total RTP Streams:', total_streams),
        ('Average MOS:', avg_mos_all),
        ('Quality Rating:', get_quality_rating(avg_mos_all)[0]),
        ('Clipped Audio Calls:', clipped_calls),  # NEW in v1.1
        ('Total Issues:', total_issues),
        ('Critical Issues:', critical_issues)
    ]
    
    row = 6
    for label, value in stats:
        ws6[f'A{row}'] = label
        ws6[f'B{row}'] = value
        ws6[f'A{row}'].font = Font(bold=True)
        if label == 'Critical Issues:' and value > 0:
            ws6[f'B{row}'].fill = PatternFill(start_color=COLOR_CRITICAL, end_color=COLOR_CRITICAL, fill_type='solid')
            ws6[f'B{row}'].font = Font(bold=True, color='FFFFFFFF')
        if label == 'Clipped Audio Calls:' and value > 0:  # NEW in v1.1
            ws6[f'B{row}'].fill = PatternFill(start_color=COLOR_POOR, end_color=COLOR_POOR, fill_type='solid')
            ws6[f'B{row}'].font = Font(bold=True)
        row += 1
    
    # MOS interpretation guide
    ws6[f'A{row+1}'] = 'MOS Score Interpretation Guide'
    ws6[f'A{row+1}'].font = Font(size=14, bold=True)
    row += 2
    
    guide = [
        ('MOS Range', 'Quality', 'Description'),
        ('4.3 - 4.5', 'Excellent', 'Crystal clear, no issues'),
        ('4.0 - 4.2', 'Good', 'Minor impairments, acceptable'),
        ('3.0 - 3.9', 'Fair', 'Noticeable quality issues'),
        ('1.0 - 2.9', 'Poor', 'Severe quality problems')
    ]
    
    for g in guide:
        ws6.append(list(g))
        for col_num in range(1, 4):
            cell = ws6.cell(row=ws6.max_row, column=col_num)
            if g == guide[0]:
                apply_cell_style(cell, cell.value, is_header=True)
            else:
                cell.border = thin_border
    
    # NEW in v1.1: Clipped Audio guide
    row = ws6.max_row + 3
    ws6[f'A{row}'] = 'Clipped Audio Detection (NEW in v1.1)'
    ws6[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    clipped_guide = [
        ('Delay Range', 'Severity', 'User Experience'),
        ('< 500ms', 'Normal', 'No noticeable delay'),
        ('500ms - 2s', 'Warning', 'First word may be cut off'),
        ('> 2s', 'Critical', '"Hello? HELLO?!" - significant clipping')
    ]
    
    for g in clipped_guide:
        ws6.append(list(g))
        for col_num in range(1, 4):
            cell = ws6.cell(row=ws6.max_row, column=col_num)
            if g == clipped_guide[0]:
                apply_cell_style(cell, cell.value, is_header=True)
            else:
                cell.border = thin_border
    
    # Issue severity guide
    row = ws6.max_row + 3
    ws6[f'A{row}'] = 'Issue Severity Levels'
    ws6[f'A{row}'].font = Font(size=14, bold=True)
    row += 1
    
    severity_guide = [
        ('Severity', 'Impact', 'Action Required'),
        ('CRITICAL', 'Call fails or unusable', 'Immediate action required'),
        ('HIGH', 'Significant degradation', 'Fix within 24 hours'),
        ('MEDIUM', 'Noticeable but usable', 'Fix within 1 week'),
        ('LOW', 'Minor optimization', 'Fix when convenient')
    ]
    
    for g in severity_guide:
        ws6.append(list(g))
        for col_num in range(1, 4):
            cell = ws6.cell(row=ws6.max_row, column=col_num)
            if g == severity_guide[0]:
                apply_cell_style(cell, cell.value, is_header=True)
            else:
                cell.border = thin_border
                if col_num == 1 and g[0] in [SEVERITY_CRITICAL, SEVERITY_HIGH, SEVERITY_MEDIUM, SEVERITY_LOW]:
                    apply_cell_style(cell, cell.value, severity=g[0])
    
    for col in range(1, 5):
        ws6.column_dimensions[get_column_letter(col)].width = 25
    
    wb.save(out_name)
    return out_name

# ==================== MAIN ====================

def main():
    print_banner()
    
    base = os.getcwd()
    pcaps = find_pcaps(base)
    
    if not pcaps:
        print(f"\n❌ No PCAP files found in {base}")
        print("Place your .pcap or .pcapng files in this directory and run again.\n")
        return
    
    print(f"\n📁 Found {len(pcaps)} PCAP file(s)\n")
    
    for pcap_path in pcaps:
        try:
            calls, rtp_streams, issues = analyze_pcap(pcap_path)
            out_file = write_excel_report(calls, rtp_streams, issues)
            
            print(f"\n{'='*70}")
            print(f"✅ Report saved: {out_file}")
            print(f"{'='*70}")
            
            # Print issue summary
            if issues:
                print(f"\n⚠️  Issue Summary:")
                critical = len([i for i in issues if i['Severity'] == SEVERITY_CRITICAL])
                high = len([i for i in issues if i['Severity'] == SEVERITY_HIGH])
                medium = len([i for i in issues if i['Severity'] == SEVERITY_MEDIUM])
                low = len([i for i in issues if i['Severity'] == SEVERITY_LOW])
                
                if critical > 0:
                    print(f"   🔴 CRITICAL: {critical}")
                if high > 0:
                    print(f"   🟠 HIGH: {high}")
                if medium > 0:
                    print(f"   🟡 MEDIUM: {medium}")
                if low > 0:
                    print(f"   🟢 LOW: {low}")
            else:
                print(f"\n✅ No issues detected - All calls are healthy!")
            
            print()
            
        except Exception as e:
            print(f"\n❌ Error processing {pcap_path}")
            traceback.print_exc()
    
    print(f"\n{'='*70}")
    print("Thank you for using VoIPScope v1.1!")
    print("NEW: Clipped Audio Detection")
    print("NEW: Silent Call Detection")
    print("NEW: Call Setup Delay Analysis")
    print("NEW: Codec Mismatch Warnings")
    print("NEW: DTMF Detection (RFC2833)")
    print("For updates and support: https://github.com/emrekarayazgan/voipscope")
    print(f"{'='*70}\n")

if __name__ == '__main__':
    main()